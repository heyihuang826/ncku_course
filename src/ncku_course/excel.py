from typing import List
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.fills import PatternFill
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.formatting import Rule
from .const import first_row



def to_excel(data : List) -> Workbook:
    columns = first_row
    print(f"共有 {len(data) - 1} 筆課程資料")
    print("正在儲存資料...")
    wb = Workbook()
    #wb = load_workbook('test.xlsx')
    ws = wb.active
    ws.append(columns)
    for r in data:
        ws.append(r)
        
    print("設定格式...")

    ws.title = "課程資料1"

    tab = Table(displayName="課程資料表格1", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")

    style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    ws.add_table(tab)

    #sheet3
    ws3 = wb.create_sheet("課表填寫")

    ws3['A1'] = '請填寫各時段可排課狀態，上方為星期，左方為節次。\n\n下表填寫狀態將會更新至「課程資料2」中「衝堂時段總數」欄位，表示該門課與本表「非空堂」時段衝堂數。'
    ws3.merge_cells('A1:H1')
    ws3['A1'].alignment = Alignment(wrap_text=True)
    ws3['A1'].font = Font(size=12)
    ws3['I2'] = '空堂(可排課)'
    ws3['J2'] = 0
    ws3['J2'].fill = PatternFill("solid", start_color="95B3D7")
    ws3['I3'] = '非空堂(不可排課)'
    ws3['J3'] = 1
    ws3['J3'].fill = PatternFill("solid", start_color="F79646")
    ws3['A2'] = "請填寫"
    ws3.column_dimensions['I'].width = 18
    ws3.column_dimensions['J'].width = 8.5
    ws3.row_dimensions[1].height = 65

    cell_range = ws3['B2:H2'][0]
    for i in range(len(cell_range)):
            cell_range[i].value = str(i + 1)

    data_list = [0, 1, 2, 3, 4, 'N', 5, 6, 7, 8, 9, 'A', 'B', 'C', 'D', 'E']
    cell_range = ws3['A3:A18']
    for i in range(len(cell_range)):
            cell_range[i][0].value = data_list[i]
            
    tab = Table(displayName="課表1", ref="A2:H18")

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    ws3.add_table(tab)
        
    for row in ws3['B3:H18']:
        for cell in row:
            cell.value = 0

    dv = DataValidation(type="list", formula1='"0, 1"', allow_blank=True, showErrorMessage = True)
    dv.error ='請輸入0或1，謝謝'
    dv.errorTitle = '輸入錯誤'
    ws3.add_data_validation(dv)

    for row in ws3['B3:H18']:
        for cell in row:
            dv.add(cell)

    data_list = ['1-0', '1-1', '1-2', '1-3', '1-4', '1-N', '1-5', '1-6', '1-7', '1-8', '1-9', '1-A', '1-B', '1-C', '1-D', '1-E', '2-0', '2-1', '2-2', '2-3', '2-4', '2-N', '2-5', '2-6', '2-7', '2-8', '2-9', '2-A', '2-B', '2-C', '2-D', '2-E', '3-0', '3-1', '3-2', '3-3', '3-4', '3-N', '3-5', '3-6', '3-7', '3-8', '3-9', '3-A', '3-B', '3-C', '3-D', '3-E', '4-0', '4-1', '4-2', '4-3', '4-4', '4-N', '4-5', '4-6', '4-7', '4-8', '4-9', '4-A', '4-B', '4-C', '4-D', '4-E', '5-0', '5-1', '5-2', '5-3', '5-4', '5-N', '5-5', '5-6', '5-7', '5-8', '5-9', '5-A', '5-B', '5-C', '5-D', '5-E', '6-0', '6-1', '6-2', '6-3', '6-4', '6-N', '6-5', '6-6', '6-7', '6-8', '6-9', '6-A', '6-B', '6-C', '6-D', '6-E', '7-0', '7-1', '7-2', '7-3', '7-4', '7-N', '7-5', '7-6', '7-7', '7-8', '7-9', '7-A', '7-B', '7-C', '7-D', '7-E']
    idx = 0
    for row in ws3['A19:DH19']:
        for cell in row:
            cell.value = data_list[idx]
            idx += 1

    data = ['=B3', '=B4', '=B5', '=B6', '=B7', '=B8', '=B9', '=B10', '=B11', '=B12', '=B13', '=B14', '=B15', '=B16', '=B17', '=B18', '=C3', '=C4', '=C5', '=C6', '=C7', '=C8', '=C9', '=C10', '=C11', '=C12', '=C13', '=C14', '=C15', '=C16', '=C17', '=C18', '=D3', '=D4', '=D5', '=D6', '=D7', '=D8', '=D9', '=D10', '=D11', '=D12', '=D13', '=D14', '=D15', '=D16', '=D17', '=D18', '=E3', '=E4', '=E5', '=E6', '=E7', '=E8', '=E9', '=E10', '=E11', '=E12', '=E13', '=E14', '=E15', '=E16', '=E17', '=E18', '=F3', '=F4', '=F5', '=F6', '=F7', '=F8', '=F9', '=F10', '=F11', '=F12', '=F13', '=F14', '=F15', '=F16', '=F17', '=F18', '=G3', '=G4', '=G5', '=G6', '=G7', '=G8', '=G9', '=G10', '=G11', '=G12', '=G13', '=G14', '=G15', '=G16', '=G17', '=G18', '=H3', '=H4', '=H5', '=H6', '=H7', '=H8', '=H9', '=H10', '=H11', '=H12', '=H13', '=H14', '=H15', '=H16', '=H17', '=H18']
    idx = 0
    for row in ws3['A20:DH20']:
        for cell in row:
            cell.value = data[idx]
            idx += 1
            
    tab = Table(displayName="課表2", ref="A19:DH20")

    ws3.add_table(tab)

    dxf = DifferentialStyle(fill=PatternFill(start_color='F79646', end_color='F79646', fill_type="solid"))
    rule = Rule(type='cellIs', dxf=dxf, operator="equal", formula=[1])
    ws3.conditional_formatting.add("B3:H18", rule)

    ws3.row_dimensions[19].hidden = True
    ws3.row_dimensions[20].hidden = True

    #sheet2
    ws_copy = wb.copy_worksheet(ws)

    ws_copy.title = "課程資料2"

    for cell in ws_copy[1]:
        cell.value = str(cell.value)

    ws_copy[f"{get_column_letter(ws_copy.max_column + 1)}1"] = "衝堂時段總數"
        
    tab = Table(displayName="課程資料表格2", ref=f"A1:{get_column_letter(ws_copy.max_column)}{ws_copy.max_row}")

    style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    ws_copy.add_table(tab)

    text_b = get_column_letter(ws_copy.max_column-112)
    text_e = get_column_letter(ws_copy.max_column-1)
    num_c = ws_copy.max_column
    for i in range(2, ws_copy.max_row + 2):
        ws_copy.cell(row = i, column = num_c).value = f"=SUMPRODUCT(${text_b}{i}:${text_e}{i},課表填寫!$A$20:$DH$20)"

    ws_copy.column_dimensions.group(get_column_letter(ws_copy.max_column - 112), get_column_letter(ws_copy.max_column - 1), hidden = True)

    #更改工作表順序
    wb._sheets[1], wb._sheets[2] = wb._sheets[2], wb._sheets[1]
    
    return wb